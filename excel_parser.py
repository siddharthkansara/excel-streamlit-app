# excel_parser.py
import numpy as np
import pandas as pd
import openpyxl
from datetime import datetime
from io import BytesIO

def process_excel(file):
    """
    Process uploaded Excel file and return a BytesIO Excel file (with multiple sheets).
    file: file-like object or path
    """
    all_data = []

    # Load workbook
    try:
        wb = openpyxl.load_workbook(file, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {file}")
    except Exception as e:
        # fallback without data_only
        wb = openpyxl.load_workbook(file, data_only=False)

    processed_sheets = 0
    for sheet_name in wb.sheetnames:
        if not (len(sheet_name) >= 5 and sheet_name[:3].isalpha() and sheet_name[3:].isdigit()):
            continue

        sheet = wb[sheet_name]
        if sheet.max_row <= 1:
            continue

        max_row = sheet.max_row
        block_num = 0
        processed_blocks = 0

        while True:  # Loop through blocks
            block_start_row = 1 + (block_num * 62)
            if block_start_row > max_row:
                break

            date_row = block_start_row + 1
            potential_date_value = None
            is_block_potentially_valid = False
            if date_row <= max_row:
                try:
                    potential_date_cell = sheet.cell(row=date_row, column=21)  # Col U
                    potential_date_value = potential_date_cell.value
                    if (potential_date_value and isinstance(potential_date_value, str) and potential_date_value.strip().upper().startswith("DATE :")) or \
                            isinstance(potential_date_value, datetime):
                        is_block_potentially_valid = True
                except Exception:
                    pass
            if not is_block_potentially_valid:
                break

            block_date = None
            try:
                if potential_date_value and isinstance(potential_date_value, str) and potential_date_value.strip().upper().startswith("DATE :"):
                    date_part = potential_date_value.split(":", 1)[-1].strip()
                    block_date = datetime.strptime(date_part, '%d.%m.%Y').date()
                elif isinstance(potential_date_value, datetime):
                    block_date = potential_date_value.date()
            except Exception:
                block_date = None

            for relative_row in range(8, 15):  # 8 to 14 inclusive
                absolute_row = block_start_row + relative_row - 1
                if absolute_row > max_row:
                    break

                try:
                    machine = sheet.cell(row=absolute_row, column=2).value  # Col B
                    # Day Shift Data
                    operator_d = sheet.cell(row=absolute_row, column=3).value  # Col C
                    roller_d = sheet.cell(row=absolute_row, column=5).value  # Col E
                    weight_d = sheet.cell(row=absolute_row, column=6).value  # Col F
                    quantity_d = sheet.cell(row=absolute_row, column=7).value  # Col G
                    remarks_d = sheet.cell(row=absolute_row, column=9).value  # Col I
                    # Night Shift Data
                    operator_n = sheet.cell(row=absolute_row, column=11).value  # Col K
                    roller_n = sheet.cell(row=absolute_row, column=15).value  # Col O
                    weight_n = sheet.cell(row=absolute_row, column=16).value  # Col P
                    quantity_n = sheet.cell(row=absolute_row, column=17).value  # Col Q
                    remarks_n = sheet.cell(row=absolute_row, column=18).value  # Col R

                    operator_d_processed = "NO NAME" if operator_d is None or str(operator_d).strip() == "" else str(operator_d).strip()
                    remarks_d_processed = "NR" if remarks_d is None or str(remarks_d).strip() == "" else str(remarks_d).strip()
                    roller_d_str = str(roller_d) if pd.notnull(roller_d) else ''

                    operator_n_processed = "NO NAME" if operator_n is None or str(operator_n).strip() == "" else str(operator_n).strip()
                    remarks_n_processed = "NR" if remarks_n is None or str(remarks_n).strip() == "" else str(remarks_n).strip()
                    roller_n_str = str(roller_n) if pd.notnull(roller_n) else ''

                    if machine is not None and str(machine).strip() != "":
                        all_data.append({
                            'Sheet': sheet_name,
                            'Date': block_date,
                            'Machine': str(machine).strip(),
                            'Shift': 'Day',
                            'Operator': operator_d_processed,
                            'RollerSize': roller_d_str,
                            'Weight': weight_d,
                            'Quantity': quantity_d,
                            'Remarks': remarks_d_processed
                        })
                        all_data.append({
                            'Sheet': sheet_name,
                            'Date': block_date,
                            'Machine': str(machine).strip(),
                            'Shift': 'Night',
                            'Operator': operator_n_processed,
                            'RollerSize': roller_n_str,
                            'Weight': weight_n,
                            'Quantity': quantity_n,
                            'Remarks': remarks_n_processed
                        })
                except Exception:
                    continue

            processed_blocks += 1
            block_num += 1

        processed_sheets += 1

    if not all_data:
        raise ValueError("No data extracted from workbook.")

    df_extracted = pd.DataFrame(all_data)

    # Conversions
    df_extracted['Date'] = pd.to_datetime(df_extracted['Date'], errors='coerce', dayfirst=True).dt.date
    df_extracted['Weight'] = pd.to_numeric(df_extracted['Weight'], errors='coerce')
    df_extracted['Quantity'] = pd.to_numeric(df_extracted['Quantity'], errors='coerce')
    for col in ['RollerSize','Operator','Machine','Remarks','Shift','Sheet']:
        df_extracted[col] = df_extracted[col].astype(str)

    target_map = {
        'HD-20': 15600, 'HD-16': 17499, 'HD-14': 19440,
        'HD-13': 21027, 'HD-12': 22032, 'HD-09': 32076,
        'HD-08': 32076
    }
    rs100_target_map = {
        'HD-20': 23600, 'HD-16': 23000, 'HD-14': 30000,
        'HD-13': 30000, 'HD-12': 33000, 'HD-09': 45000,
        'HD-08': 45000
    }
    df_extracted['Target'] = pd.to_numeric(df_extracted['Machine'].map(target_map), errors='coerce')
    df_extracted['Rs 100 Target'] = pd.to_numeric(df_extracted['Machine'].map(rs100_target_map), errors='coerce')
    df_extracted['Rs 100 Bonus'] = np.where(
        (df_extracted['Quantity'].notna()) & (df_extracted['Rs 100 Target'].notna()) & (df_extracted['Quantity'] > df_extracted['Rs 100 Target']),
        100,
        0
    )

    # Fill zeros for grouping
    df_extracted['Quantity'] = df_extracted['Quantity'].fillna(0)
    df_extracted['Target'] = df_extracted['Target'].fillna(0)
    df_extracted['Rs 100 Target'] = df_extracted['Rs 100 Target'].fillna(0)

    # Machine summary
    summary_with_grand_total = pd.DataFrame()
    if not df_extracted.empty:
        summary_by_machine = df_extracted.groupby('Machine').agg(
            Total_Quantity=('Quantity', 'sum'),
            Total_Target=('Target', 'sum'),
            Total_Rs_100_Target=('Rs 100 Target', 'sum')
        ).reset_index()
        grand_total_row_machine = pd.DataFrame({
            'Machine': ['Grand Total'],
            'Total_Quantity': [summary_by_machine['Total_Quantity'].sum()],
            'Total_Target': [summary_by_machine['Total_Target'].sum()],
            'Total_Rs_100_Target': [summary_by_machine['Total_Rs_100_Target'].sum()]
        })
        summary_with_grand_total = pd.concat([summary_by_machine, grand_total_row_machine], ignore_index=True)

    # Roller summary
    roller_summary_df = pd.DataFrame()
    if not df_extracted.empty and 'RollerSize' in df_extracted.columns:
        roller_summary_df = df_extracted.groupby('RollerSize').agg(
            Total_Quantity_By_Roller=('Quantity', 'sum')
        ).reset_index()

    # Write to Excel in memory
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_extracted.to_excel(writer, sheet_name='Detailed_Data', index=False)
        if not summary_with_grand_total.empty:
            summary_with_grand_total.to_excel(writer, sheet_name='Machine_Summary', index=False)
        if not roller_summary_df.empty:
            roller_summary_df.to_excel(writer, sheet_name='RollerSize_Summary', index=False)
    output.seek(0)
    return output
